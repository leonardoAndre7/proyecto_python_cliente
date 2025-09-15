import os
from django.http import HttpResponse
from decimal import Decimal
import pandas as pd
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.views.decorators.csrf import csrf_exempt
from django.shortcuts import render, redirect
from django.conf import settings
from django.db import transaction
from .forms import UploadExcelForm
from .models import BCP, TarifaOperacion, Cliente
from django.urls import reverse
import re
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from decimal import Decimal
import pandas as pd
from .models import Cliente, TarifaOperacion



def importar_excel(request):
    """
    Función que permite subir un archivo Excel y mostrar una PREVIEW editable
    antes de guardar los datos en la base de datos.
    """
    if request.method == "POST":
        form = UploadExcelForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            cliente_default = form.cleaned_data.get('cliente')
            tarifa_default = form.cleaned_data.get('tarifa')
            saldo_inicial_default = form.cleaned_data.get('saldo_inicial') or Decimal('0.00')

            df = pd.read_excel(archivo, sheet_name=0, dtype=str)
            df.columns = [c.strip().upper().replace(" ", "_") for c in df.columns]

            rows = []

            for i, r in df.iterrows():
                cod_bcp = str(r.get("COD_BCP") or r.get("COD_BCP ") or "")

                # --- Fechas manejando valores nulos ---
                fecha_raw = r.get("FECHA")
                fecha_valuta_raw = r.get("FECHA_VALUTA") or r.get("FECHA_VAL") or None

                def safe_fecha(f):
                    try:
                        if pd.isna(f):
                            return None
                        return pd.to_datetime(f)
                    except:
                        return None

                fecha = safe_fecha(fecha_raw)
                fecha_valuta = safe_fecha(fecha_valuta_raw)

                # Otros campos
                descripcion = r.get("DESCRIPCIÓN_OPERACIÓN") or r.get("DESCRIPCION_OPERACION") or ""
                sucursal_agencia = r.get("SUCURSAL_AGENCIA") or ""
                n_operacion = r.get("N_OPERACIÓN") or r.get("N_OPERACION") or ""

                # Monto
                monto_raw = r.get("MONTO") or r.get("MTO") or "0"
                try:
                    monto = Decimal(str(monto_raw).replace(",", ""))
                except:
                    monto = Decimal('0.00')

                # Extraer DNI de los últimos 8 dígitos de la descripción
                match = re.search(r'(\d{8})$', descripcion)
                dni_encontrado = match.group(1) if match else None

                # Inicializamos datos del cliente
                cliente_obj = None
                if dni_encontrado:
                    cliente_obj = Cliente.objects.filter(dni=dni_encontrado).first()

                if not cliente_obj and cliente_default:
                    cliente_obj = cliente_default

                # Inicializamos datos de tarifa
                # Inicializamos datos de tarifa
                tarifa_obj = None
                if cliente_obj and cliente_obj.cod_tarifa:
                    tarifa_obj = TarifaOperacion.objects.filter(cod_tarifa=cliente_obj.cod_tarifa).first()

                if not tarifa_obj and tarifa_default:
                    tarifa_obj = tarifa_default


                # Crear objeto temporal BCP para cálculos
                bcp_temp = BCP(
                monto=monto,
                cliente=cliente_obj,
                tarifa=tarifa_obj,
            )


                # Asignar saldo inicial al atributo temporal
                bcp_temp._saldo_inicial = saldo_inicial_default

                # Calcular saldo, comisión y lm_pagar
                bcp_temp.calcular_datos()

                # Diccionario de fila para preview
                row = {
                    "index": i,
                    "cod_bcp": cod_bcp,
                    "fecha": '' if fecha is None else fecha.strftime('%Y-%m-%d'),
                    "fecha_valuta": '' if fecha_valuta is None else fecha_valuta.strftime('%Y-%m-%d'),
                    "descripcion": descripcion,
                    "monto": monto,
                    "sucursal_agencia": sucursal_agencia,
                    "n_operacion": n_operacion,
                    "usuario": r.get("USUARIO") or "",
                    # Datos automáticos del cliente
                    "dni_cliente": dni_encontrado,
                    "cliente_default": cliente_obj.cod_cliente if cliente_obj else "",
                    "cliente_nombre": cliente_obj.nombre if cliente_obj else "",
                    "cliente_apellidos": cliente_obj.apellidos if cliente_obj else "",
                    "correo": cliente_obj.correo or "" if cliente_obj else "",
                    "celular": cliente_obj.celular or "" if cliente_obj else "",
                    "status": cliente_obj.status or "" if cliente_obj else "",
                    "provincia": cliente_obj.provincia or "" if cliente_obj else "",
                    "codigo_referido": cliente_obj.codigo_referido or "" if cliente_obj else "",
                    "nombre_referido": cliente_obj.nombre_referido or "" if cliente_obj else "",
                    "cuenta_banco_referido": cliente_obj.cuenta_banco_referido or "" if cliente_obj else "",
                    "cuenta_interbancario_referido": cliente_obj.cuenta_interbancario_referido or "" if cliente_obj else "",
                    #faltan esos 2
                    "costo_por_porcentaje": tarifa_obj.costo_por_porcentaje if tarifa_obj else 0,
                    "costo_fijo": tarifa_obj.costo_fijo if tarifa_obj else 0,
                    #faltan los 2 de arriba
                    # Datos calculados automáticamente
                    "saldo": f"{bcp_temp.saldo:.2f}",
                    "comision": f"{bcp_temp.comision:.2f}",
                    "lm_pagar": f"{bcp_temp.lm_pagar:.2f}",
                    "tarifa_default": tarifa_obj.cod_tarifa if tarifa_obj else "",
                    "saldo_inicial_default": saldo_inicial_default,
                    "Ganancia_Referido": f"{bcp_temp.ganancia_referido:.2f}",

                    "cod_tarifa": cliente_obj.cod_tarifa or "" if cliente_obj else "",
                }

                rows.append(row)

            # Contexto para renderizar el template
            context = {
                "rows": rows,
                "clientes": Cliente.objects.all(),
                "tarifas": TarifaOperacion.objects.all(),
                "form": form,
            }

            return render(request, "banco/preview.html", context)

    else:
        form = UploadExcelForm()

    return render(request, "banco/importar_excel.html", {"form": form})


def exportar_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BCP"

    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    money_format = "#,##0.00"
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Encabezados
    headers = [
        "Código BCP", "Fecha", "Fecha Valuta", "Descripción", 
        "Monto", "Sucursal/Agencia", "N° Operación", "Usuario",
        "Saldo", "Comisión", "LM Pagar", "Código", "Ganancia Referido",
        "Cliente", "Tarifa"
    ]
    ws.append(headers)

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = thin_border

    # Filas de datos
    for i, bcp in enumerate(BCP.objects.all(), start=2):
        row = [
            bcp.cod_bcp,
            bcp.fecha,
            bcp.fecha_valuta,
            bcp.descripcion,
            float(bcp.monto or 0),
            bcp.sucursal_agencia,
            bcp.n_operacion,
            bcp.usuario,
            float(bcp.saldo or 0),
            float(bcp.comision or 0),
            float(bcp.lm_pagar or 0),
            bcp.codigo,
            float(bcp.ganancia_referido or 0),
            bcp.cliente.cod_cliente if bcp.cliente else "",
            bcp.tarifa.cod_tarifa if bcp.tarifa else ""
        ]
        ws.append(row)

        # Aplicar bordes y formato monetario
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=i, column=col_num)
            cell.border = thin_border
            if col_num in [5, 9, 10, 11, 13]:  # columnas de dinero
                cell.number_format = money_format

        # Efecto zebra
        if i % 2 == 0:
            for col_num in range(1, len(row)+1):
                ws.cell(row=i, column=col_num).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # Ajuste automático de ancho
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = max_length

    # Respuesta HTTP
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="BCP.xlsx"'
    wb.save(response)
    return response

@transaction.atomic
def confirmar_import(request):
    """
    Función que guarda los registros confirmados desde la preview.
    """
    if request.method != "POST":
        return redirect(reverse("banco:importar_excel"))

    # Obtenemos listas de cada campo
    cod_bcp_list = request.POST.getlist("cod_bcp")
    fecha_list = request.POST.getlist("fecha")
    fecha_valuta_list = request.POST.getlist("fecha_valuta")
    descripcion_list = request.POST.getlist("descripcion")
    monto_list = request.POST.getlist("monto")
    sucursal_list = request.POST.getlist("sucursal_agencia")
    n_operacion_list = request.POST.getlist("n_operacion")
    usuario_list = request.POST.getlist("usuario")
    saldo_inicial_list = request.POST.getlist("saldo_inicial")
    codigo_list = request.POST.getlist("codigo")
    cliente_list = request.POST.getlist("cliente")
    tarifa_list = request.POST.getlist("tarifa")

    saved = 0
    errors = []
    total = len(descripcion_list)

    for i in range(total):
        try:
            cod_bcp = cod_bcp_list[i] if i < len(cod_bcp_list) else None

            # Fechas
            f_raw = fecha_list[i] if i < len(fecha_list) else ""
            try:
                fecha = pd.to_datetime(f_raw).date() if f_raw else None
            except:
                fecha = None

            fv_raw = fecha_valuta_list[i] if i < len(fecha_valuta_list) else ""
            try:
                fecha_valuta = pd.to_datetime(fv_raw).date() if fv_raw else None
            except:
                fecha_valuta = None

            descripcion = descripcion_list[i]
            monto = Decimal(monto_list[i].replace(",", "")) if monto_list[i] else Decimal('0.00')
            sucursal_agencia = sucursal_list[i] if i < len(sucursal_list) else ""
            n_operacion = n_operacion_list[i] if i < len(n_operacion_list) else ""
            usuario = usuario_list[i] if i < len(usuario_list) else ""
            saldo_inicial = Decimal(saldo_inicial_list[i].replace(",", "")) if saldo_inicial_list[i] else Decimal('0.00')
            codigo = codigo_list[i] if i < len(codigo_list) else ""

            cliente_obj = Cliente.objects.filter(cod_cliente=cliente_list[i]).first() if cliente_list and cliente_list[i] else None
            tarifa_obj = TarifaOperacion.objects.filter(cod_tarifa=tarifa_list[i]).first() if tarifa_list and tarifa_list[i] else None

            if cod_bcp and BCP.objects.filter(cod_bcp=cod_bcp).exists():
                continue

            b = BCP(
                cod_bcp=cod_bcp or f"autogen_{i}",
                fecha=fecha,
                fecha_valuta=fecha_valuta,
                descripcion=descripcion,
                monto=monto,
                sucursal_agencia=sucursal_agencia,
                n_operacion=n_operacion,
                usuario=usuario,
                codigo=codigo,
                cliente=cliente_obj,
                tarifa=tarifa_obj
            )

            b.saldo_inicial = saldo_inicial
            b.save()
            saved += 1
        except Exception as e:
            errors.append(f"Fila {i}: {str(e)}")

    return render(request, "banco/result.html", {"saved": saved, "errors": errors})
